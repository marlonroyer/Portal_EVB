import json
from pathlib import Path

import openpyxl


PROJECT_DIR = Path(__file__).resolve().parents[1]
SOURCE = PROJECT_DIR / "exemplos" / "Mapa_Coleção_Inventário_B04_E02_S11.xlsx"
OUTPUT = PROJECT_DIR / "data" / "collection-map.js"


def clean(value):
    return str(value).strip() if value is not None else ""


def split_lines(value, keep_blank=False):
    lines = [item.strip() for item in clean(value).split("\n")]
    return lines if keep_blank else [item for item in lines if item]


def get_space_count(sheet_name):
    cabinet_number = int("".join(char for char in sheet_name if char.isdigit()) or "0")
    return 2 if cabinet_number >= 6 else 4


def build_collection_map():
    workbook = openpyxl.load_workbook(SOURCE, data_only=True)
    cabinets = []
    boxes = []
    total_spaces = 0
    occupied_spaces = 0
    free_spaces = 0

    for sheet_name in [f"A{index}" for index in range(1, 9)]:
        worksheet = workbook[sheet_name]
        shelves = []

        for location_row in range(3, 14, 2):
            content_row = location_row - 1
            shelf_id = clean(worksheet.cell(location_row, 1).value) or f"P{(location_row - 1) // 2}"
            spaces = []

            for space_index in range(1, get_space_count(sheet_name) + 1):
                start_column = 2 + ((space_index - 1) * 5)
                space_id = f"E{space_index}"
                location = f"{sheet_name}-{shelf_id}-{space_id}"
                box_id = clean(worksheet.cell(content_row, start_column).value)
                families = split_lines(worksheet.cell(content_row, start_column + 1).value, keep_blank=True)
                content = split_lines(worksheet.cell(content_row, start_column + 2).value)
                notes = clean(worksheet.cell(content_row, start_column + 3).value)
                free_marker = any(
                    clean(worksheet.cell(content_row, start_column + offset).value).lower() == "livre"
                    for offset in range(4)
                )
                has_location = (
                    clean(worksheet.cell(location_row, start_column + 1).value)
                    or clean(worksheet.cell(location_row, start_column + 2).value)
                )

                if not any([box_id, families, content, notes, free_marker, has_location]):
                    continue

                status = "occupied" if box_id else "free"
                record = {
                    "id": location,
                    "cabinet": sheet_name,
                    "shelf": shelf_id,
                    "space": space_id,
                    "location": location,
                    "box": box_id,
                    "families": families,
                    "content": content,
                    "notes": notes,
                    "status": status,
                    "source": {
                        "sheet": sheet_name,
                        "contentRow": content_row,
                        "locationRow": location_row,
                        "startColumn": start_column,
                    },
                }

                spaces.append(record)
                total_spaces += 1
                if status == "occupied":
                    occupied_spaces += 1
                    boxes.append(record)
                else:
                    free_spaces += 1

            shelves.append({"id": shelf_id, "spaces": spaces})

        cabinets.append({"id": sheet_name, "shelves": shelves})

    return {
        "sourceFile": SOURCE.name,
        "generatedAt": "2026-07-01",
        "notes": "Dados derivados das abas A1-A8. Localizacoes da aba A8 foram normalizadas para A8 quando havia residuos de copia da A7.",
        "stats": {
            "cabinets": len(cabinets),
            "spaces": total_spaces,
            "occupied": occupied_spaces,
            "free": free_spaces,
            "boxes": len(boxes),
        },
        "cabinets": cabinets,
        "boxes": boxes,
    }


def main():
    data = build_collection_map()
    OUTPUT.write_text(
        "window.EVB_COLLECTION_MAP = "
        + json.dumps(data, ensure_ascii=True, indent=2)
        + ";\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
